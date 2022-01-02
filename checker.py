#!/usr/bin/env python3

import re, argparse, sys, json, logging, os, util
from openpyxl import load_workbook
from numbers import Number
from datetime import datetime
from violations import NonEmptyViolation, TypeViolation, RegexViolation
from tabulate import tabulate

"""
    This dictionary will be used to check if the values
    of a column have the expected type.
"""
TYPES = {
    "string" : str,
    "number" : Number,
    "date" : datetime
}

# Regex will only be checked for this type
STRING_TYPE = "string"

# Maximum number of rows in the output tables
OUTPUT_TABLE_MAX = 20

# Only execute if run as a script
if __name__ == "__main__":

    # Argument parsing
    parser = argparse.ArgumentParser()
    parser.add_argument("table", help=".xlsx file that needs to be checked.", type=str)
    parser.add_argument("structure", help=".json structure file", type=str)
    parser.add_argument("-s, --sheet", dest="sheet", help="Name of the sheet", type=str)
    parser.add_argument("-o, --highlighted", dest="highlighted", help="Store copy of .xlsx file with violating cells highlighted.", type=str)
    parser.add_argument("--hide-skipped", dest="hide_skipped", action="store_const", const=True, help="Dont print skipped columns")
    parser.add_argument("--hide-ok", dest="hide_ok", action="store_const", const=True, help="Dont print columns without violations")
    args = parser.parse_args()

    table_filename  = args.table
    struct_filename = args.structure
    sheetname = args.sheet
    hide_skipped = args.hide_skipped
    hide_ok = args.hide_ok
    hl_filename = args.highlighted

    # Logging
    logging.basicConfig(format='%(levelname)s: %(message)s', level=logging.INFO)

    # Check file endings
    if not table_filename.endswith(".xlsx"):
        logging.critical("Table file must be of type '.xlsx' .")
        sys.exit()
    if not struct_filename.endswith(".json"):
        logging.critical("Structure file must be of type '.json' .")
        sys.exit()
    if hl_filename != None and not hl_filename.endswith(".xlsx"):
        logging.critical("Highlighted output file must be of type '.xlsx' .")
        sys.exit()
    if hl_filename == table_filename: # Forbid overwriting original file
        logging.warning("Highlighted output file cannot be input table file.")
        sys.exit()

    # Load structure
    print(f"Loading structure file {struct_filename.split(os.path.sep)[-1]} .. ")
    with open(struct_filename, "r") as struct_file:
        structure = json.load(struct_file)
    struct_check = util.check_struct(structure)
    if struct_check != None:
        sys.exit(f"Structure is invalid: {struct_check}")
    cols = structure["cols"]

    # Load excel sheet
    print(f"Loading excel file {table_filename.split(os.path.sep)[-1]} ..")
    try:
        read_only = hl_filename == None
        wb = load_workbook(filename=table_filename, read_only=read_only)
    except Exception as e:
        logging.critical(f"Could not open table file: {e}")
        sys.exit()

    ws = wb.active if sheetname == None else wb[sheetname]

    if ws.max_row == 1 or ws.max_column == 1:
        logging.critical("Worksheet is empty.")
        sys.exit()

    row_num = ws.max_row - 1
    print(f"Loaded file with {row_num} data rows.")

    # Check if columns exists (in the specified order)
    colnames_expected = [ col["name"] for col in cols ]
    colnames_actual   = [ cell.value for cell in list(ws.rows)[0] ]

    print(f"Checking basic column structure ..   ", end="")
    for expected, actual in zip(colnames_expected, colnames_actual):
        if expected != actual:
            sys.exit(f"\nExpected column name '{expected}', got '{actual}' instead.")

    if len(colnames_actual) != len(colnames_expected):
        sys.exit(f"\nExpected {len(colnames_expected)} column(s), got {len(colnames_actual)} instead.")
    print("Done!")

    # count non-empty cells for each row
    col_lens = { col["name"] : 0 for col in cols }

    # store violations in this dict
    violations = { col["name"] : [] for col in cols }

    # Check if value for each column has the correct type
    # If the type is a string, also check if the regular expression matches
    for index, row in enumerate(ws.iter_rows(min_row=2)):
        print(f"\rChecking row {index + 1} of {row_num} ..", end="")
        for col, cell in zip(cols, row):
            value = cell.value
            colname = col["name"]
            if value != None: # Count non-empty cells for each column
                col_lens[colname] += 1

            # skip column?
            if col.get("skip", False):
                continue

            # check if cell is allowed to be empty
            if value == None:
                if col.get("non-null", False):
                    vl = NonEmptyViolation(colname, index, str(value))
                    violations[colname].append(vl)

                    if hl_filename != None: util.mark(cell, vl)
                continue

            # Get expected type (required field)
            try:
                coltype = col["type"]
            except KeyError:
                logging.warning(f"'type' key was not found for column {colname}. Column will be skipped.")
                col["skip"] = True
            expected_type = TYPES[coltype]

            # check type
            if not isinstance(cell.value, expected_type):
                vl = TypeViolation(colname, cell.row, str(value), expected_type, type(value))
                violations[colname].append(vl)

                if hl_filename != None: util.mark(cell, vl)
                continue

            # check regex
            if col["type"] == STRING_TYPE and "regex" in col:
                pattern = col["regex"]
                if not re.match(pattern, value):
                    vl = RegexViolation(colname, cell.row, str(value), pattern)
                    violations[colname].append(vl)

                    if hl_filename != None: util.mark(cell, vl)

    print("\nDone!\n")

    # Print summary of violations
    for col in cols:
        key = col["name"]
        value = violations[key]
        skip = col.get("skip", False)

        # check if skipped should be hidden
        if hide_skipped and skip:
            continue

        # check if ok should be skipped
        if hide_ok and len(value) == 0:
            continue

        print(util.b(f"> {key}"))
        if col.get("skip", False):
            print(f"{util.SKIPPED}\n")
        elif len(value) == 0:
            print(f"{util.OK} : No violations found\n")
        else:
            print(f"{util.ERROR} : {len(value)} violations found")

            """
                For each type of violation, check if every cell in this column
                caused a violation. Print a simplified message in this case.
            """

            # type violations
            type_violations = [ vl for vl in value if isinstance(vl, TypeViolation) ]
            if len(type_violations) == col_lens[key]: # all cells
                actual_types = set([ vl.actual.__name__ for vl in type_violations ])
                util.print_indent(f"All cells did not match the expected type '{col['type']}'. Instead, the following type(s) were found: [{','.join(actual_types)}]", indent=2)

            elif len(type_violations) != 0: # some cells
                util.print_indent(f"The following cells did not match the expected type ({col['type']}) :\n", indent=2)

                table = [ (vl.row, f"'{vl.value}'", vl.actual.__name__) for vl in type_violations ]
                if len(type_violations) > OUTPUT_TABLE_MAX: # too many
                    util.print_indent(tabulate(table[:OUTPUT_TABLE_MAX], headers=("Row", "Value", "Type")), indent=4)
                    util.print_indent(f".. and {len(type_violations) - OUTPUT_TABLE_MAX} more!", indent=4)
                else:
                    util.print_indent(tabulate(table, headers=("Row", "Value", "Type")), indent=4)
            print() # padding

            # regex violations
            regex_violations = [ vl for vl in value if isinstance(vl, RegexViolation) ]
            if len(regex_violations) == col_lens[key]:
                util.print_indent(f"All cells did not match the regular expression.", indent=2)

            elif len(regex_violations) != 0:
                util.print_indent("The following cells did not match the regular expression:\n", indent=2)

                table = [ (vl.row, f"'{vl.value}'") for vl in regex_violations ]
                if len(regex_violations) > OUTPUT_TABLE_MAX: # too many
                    util.print_indent(tabulate(table[:OUTPUT_TABLE_MAX], headers=("Row", "Value")), indent=4)
                    util.print_indent(f".. and {len(regex_violations) - OUTPUT_TABLE_MAX} more!", indent=4)
                else:
                    util.print_indent(tabulate(table, headers=("Row", "Value")), indent=4)
            print() # padding

            # non empty violations
            nonempty_violations = [ vl for vl in value if isinstance(vl, NonEmptyViolation) ]
            if len(nonempty_violations) == col_lens[key]:
                util.print_indent(f"All cells are empty, even though non-null is set to true", indent=2)

            elif len(nonempty_violations) != 0:
                util.print_indent("The following cells are empty, even though non-null is set to true:\n", indent=2)

                table = [ (vl.row,) for vl in nonempty_violations ]
                if len(nonempty_violations) > OUTPUT_TABLE_MAX:
                    util.print_indent(tabulate(table[:OUTPUT_TABLE_MAX], headers=("Row",)), indent=4)
                    util.print_indent(f".. and {len(nonempty_violations) - OUTPUT_TABLE_MAX} more!", indent=4)
                else:
                    util.print_indent(tabulate(table, headers=("Row",)), indent=4)
            print() # padding


    # Save worksheet if highlight file is specified
    if hl_filename != None:
        try:
            wb.save(hl_filename)
        except PermissionError as e:
            logging.critical(f"Could not save highlight file, is the file currently open? : {e}")
